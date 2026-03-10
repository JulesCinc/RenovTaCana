"""
Exploration complète des données RenovTaCana :
- Comparaison longueur wMain vs pipe_ranking
- Schéma de tous les shapefiles (conduites, UDI, abandonnées)
- En-têtes des fichiers Excel
"""
import csv
import struct
import os
import json

# Chemins relatifs à l'emplacement du script (exécutable depuis n'importe où)
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_ROOT_DIR = os.path.dirname(os.path.dirname(_SCRIPT_DIR))
_DATA_DIR = os.path.join(_ROOT_DIR, "row-data")

def dbf_fields(path):
    """Noms des champs d'un .dbf."""
    with open(path, "rb") as f:
        data = f.read(64)
    if len(data) < 32:
        return []
    header_len = struct.unpack("<H", data[8:10])[0]
    with open(path, "rb") as f:
        f.seek(32)
        raw = f.read(header_len - 32 - 1)
    fields = []
    i = 0
    while i + 32 <= len(raw):
        name = raw[i:i+11].split(b"\x00")[0].decode("latin-1", errors="replace").strip()
        if name:
            fields.append(name)
        i += 32
    return fields

def dbf_records(path, keys):
    """Lit des champs spécifiques depuis un .dbf (nécessite dbfread)."""
    try:
        from dbfread import DBF
        table = DBF(path, encoding='latin-1')
        key_set = set(keys)
        for record in table:
            yield {k: record.get(k) for k in keys if k in record}
    except ImportError:
        return
    except Exception as e:
        print("  [Erreur dbfread]", path, e)
        return

def compare_longueur():
    """Compare longueur entre wMain et pipe_ranking (même FACILITYID)."""
    csv_path = os.path.join(_DATA_DIR, "pipe_ranking_v1_clear.csv")
    dbf_path = os.path.join(_DATA_DIR, "conduites", "wMain.dbf")

    # Charger pipe_ranking (dédoublonné : premier enregistrement par FACILITYID)
    pr_by_id = {}
    with open(csv_path, "r", encoding="utf-8") as f:
        r = csv.DictReader(f)
        for row in r:
            fid = row.get("FACILITYID")
            if fid and fid not in pr_by_id:
                try:
                    pr_by_id[fid] = float(row.get("longueur", 0) or 0)
                except (ValueError, TypeError):
                    pass

    # Charger wMain longueur depuis DBF
    wmain_by_id = {}
    for rec in dbf_records(dbf_path, ["FACILITYID", "longueur"]):
        if not rec:
            break
        fid = rec.get("FACILITYID")
        if fid:
            try:
                v = rec.get("longueur")
                wmain_by_id[fid] = float(v) if v is not None else None
            except (ValueError, TypeError):
                wmain_by_id[fid] = None

    if not wmain_by_id:
        print("  [dbfread non installé] pip install dbfread pour comparer les longueurs")
        return None

    # Comparaison
    common = set(pr_by_id) & set(wmain_by_id)
    same = 0
    diff = 0
    diffs_sample = []
    for fid in common:
        pr_val = pr_by_id[fid]
        wm_val = wmain_by_id[fid]
        if wm_val is None:
            continue
        if abs((pr_val or 0) - (wm_val or 0)) < 0.01:  # tolérance 1 cm
            same += 1
        else:
            diff += 1
            if len(diffs_sample) < 5:
                diffs_sample.append((fid, pr_val, wm_val))

    return {
        "nb_pipe_ranking": len(pr_by_id),
        "nb_wMain": len(wmain_by_id),
        "nb_common": len(common),
        "longueur_identique": same,
        "longueur_differente": diff,
        "echantillon_diffs": diffs_sample,
    }

def xlsx_headers(path):
    out = {}
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            row1 = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            out[sheet] = [str(c) if c is not None else "" for c in row1]
        wb.close()
    except Exception as e:
        out["_error"] = str(e)
    return out

def find_all_dbf():
    """Liste tous les .dbf sous le dossier données avec leur schéma (colonnes)."""
    results = []
    for root, _, files in os.walk(_DATA_DIR):
        for f in files:
            if f.lower().endswith(".dbf"):
                path = os.path.join(root, f)
                rel = os.path.relpath(path, _ROOT_DIR)
                fields = dbf_fields(path)
                name = os.path.splitext(f)[0]
                results.append((rel, name, fields))
    return results

def main():
    print("=== 1. COMPARAISON LONGUEUR wMain vs pipe_ranking ===\n")
    res = compare_longueur()
    if res:
        print("  pipe_ranking: {} lignes (FACILITYID uniques)".format(res["nb_pipe_ranking"]))
        print("  wMain:       {} enregistrements".format(res["nb_wMain"]))
        print("  En commun:   {} FACILITYID".format(res["nb_common"]))
        print("  Longueur identique (tol. 0.01 m): {}".format(res["longueur_identique"]))
        print("  Longueur différente: {}".format(res["longueur_differente"]))
        if res["echantillon_diffs"]:
            print("  Exemples de différences (FACILITYID, pipe_ranking.longueur, wMain.longueur):")
            for fid, pr, wm in res["echantillon_diffs"]:
                print("    ", fid, pr, wm)
    else:
        print("  Impossible (installer dbfread).")

    print("\n=== 2. TOUS LES SHAPEFILES (colonnes) ===\n")
    for rel, name, fields in find_all_dbf():
        print("  {}  -> {} colonnes: {}".format(rel, len(fields), fields[:8] if len(fields) > 8 else fields))
        if len(fields) > 8:
            print("       ... et", fields[8:])

    print("\n=== 3. FICHIERS EXCEL (première ligne = colonnes) ===\n")
    for xlsx in ["Operations.xlsx", "chantiers.xlsx"]:
        path = os.path.join(_DATA_DIR, xlsx)
        if os.path.exists(path):
            print("  ---", xlsx)
            for sheet, cols in xlsx_headers(path).items():
                if sheet.startswith("_"):
                    print("    Erreur:", cols)
                else:
                    print("    Feuille", repr(sheet), ":", cols)

if __name__ == "__main__":
    main()
