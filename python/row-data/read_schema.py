"""Lit les schémas (colonnes) de tous les fichiers de données pour documentation."""
import csv
import struct
import os

# Chemins relatifs à l'emplacement du script (exécutable depuis n'importe où)
_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_ROOT_DIR = os.path.dirname(os.path.dirname(_SCRIPT_DIR))  # racine du projet
_DATA_DIR = os.path.join(_ROOT_DIR, "row-data")

def dbf_fields(path):
    """Lit les noms des champs d'un fichier .dbf (format dBASE III/IV)."""
    with open(path, "rb") as f:
        data = f.read(64)
    if len(data) < 32:
        return []
    # Byte 8-9: header length (little-endian)
    header_len = struct.unpack("<H", data[8:10])[0]
    with open(path, "rb") as f:
        f.seek(32)
        raw = f.read(header_len - 32 - 1)
    fields = []
    i = 0
    while i + 32 <= len(raw):
        name = raw[i:i+11].split(b"\x00")[0].decode("latin-1", errors="replace").strip()
        if name:
            fields.append(name)
        i += 32
    return fields

def xlsx_headers(path):
    """Lit la première ligne de chaque feuille (colonnes) d'un xlsx via CSV-like read."""
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        out = {}
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            row1 = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            out[sheet] = [str(c) if c is not None else "" for c in row1]
        wb.close()
        return out
    except Exception as e:
        return {"_error": str(e)}

def main():
    # CSV
    csv_path = os.path.join(_DATA_DIR, "pipe_ranking_v1_clear.csv")
    if not os.path.exists(csv_path):
        print("Fichier non trouvé:", csv_path)
        return
    with open(csv_path, "r", encoding="utf-8") as f:
        csv_cols = next(csv.reader(f))
    print("=== pipe_ranking_v1_clear.csv ===")
    print("Colonnes:", csv_cols)

    # Tous les shapefiles (tous les .dbf sous le dossier données)
    print("\n=== Tous les shapefiles (.dbf) ===")
    for root, _, files in os.walk(_DATA_DIR):
        for f in sorted(files):
            if f.lower().endswith(".dbf"):
                dbf_path = os.path.join(root, f)
                name = os.path.splitext(f)[0]
                fields = dbf_fields(dbf_path)
                print("\n  ", dbf_path, "->", name)
                print("    Colonnes:", fields)

    # Excel
    for xlsx in ["Operations.xlsx", "chantiers.xlsx"]:
        path = os.path.join(_DATA_DIR, xlsx)
        if os.path.exists(path):
            print("\n=== " + xlsx + " ===")
            for sheet, cols in xlsx_headers(path).items():
                if sheet.startswith("_"):
                    print("  Erreur:", cols)
                else:
                    print("  Feuille", repr(sheet), ":", cols)

if __name__ == "__main__":
    main()
